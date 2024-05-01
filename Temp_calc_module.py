import numpy as np
from parapy.core import *
from m_engine_init import me
import matlab
import matlab.engine
import openpyxl
class Heating(Base):
    '''Define relavant properties for thermal model 
    and provides temperature
    todo change name of file to heating
    todo fix Input dependencies'''

    #Inputs:
    Body = Input('Mars')
    # Temperatures
    T_inside = Input(20)  # inside temp required
    T_min = Input(-150)  # Bi.cell(row=2, column=c).value  #Coldest temp on body

    # Geometry
    A_base = Input(50)  # Base Area in m2
    A_vertical = Input(560)  # Vertical Area in m2
    t_base = Input(0.25)  # Foundation Thickness in m

    # Power
    Q_sys = Input(25000)  # Nominal Power use in W
    Q_max = Input(0.05)  # percentage of heating power alowed

    @Attribute
    def get_R_info(self):
        b = self.Body
        specs = openpyxl.load_workbook('Habitat_Design_Specification.xlsx')
        Bi = specs['Environmental Input Data']

        if b == 'Mars':
            c = 3
        elif b == 'Moon':
            c = 4
        else:
            c = 5

        R_c = float(Bi.cell(row=4, column=c).value)
        R_e = float(Bi.cell(row=5, column=c).value)
        return R_c, R_e

    #'''Still have to pass the R_c and R_e in MAtlab'''
    @Attribute
    def Thermal_model(self):
        input = matlab.double([self.get_R_info[0],self.get_R_info[1],
                               self.T_inside,self.T_min,
                               self.A_base,self.A_vertical,self.t_base,
                               self.Q_sys,self.Q_max])

        x = me.Thermal_sizing(input,nargout=1)
        return np.array(x)
    @Attribute
    def Q_heat(self):
        Q_heat = self.Thermal_model[0][0]
        return Q_heat
    @Attribute
    def t_min(self):
        t_min = self.Thermal_model[0][1]
        return t_min
